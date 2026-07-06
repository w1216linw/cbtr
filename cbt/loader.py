import logging
import os
import re
import sys

import openpyxl

_log = logging.getLogger('cbt_report')

_POD_EXPECTED_HEADERS = {
    0:  '实际揽收时间',
    1:  '详细地址',
    8:  '城市',
    13: '揽收状态',
    14: '揽收失败原因',
}


def load_db(path: str) -> list:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    result = [str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    wb.close()
    return result


def load_pod(path: str):
    """Returns (filtered_rows, all_dates_in_file)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for idx, expected in _POD_EXPECTED_HEADERS.items():
        actual = headers[idx] if idx < len(headers) else None
        if actual != expected:
            wb.close()
            msg = (f'pod.xlsx 列头错误：第 {idx+1} 列期望 "{expected}"，实际是 "{actual}"'
                   f'，请检查 pod.xlsx 是否新增或删除了列')
            _log.error(msg)
            sys.exit(f'❌  {msg}')

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        reason = row[14]
        rows.append({
            'time':   str(row[0]) if row[0] else '',
            'addr':   str(row[1]).strip(),
            'city':   str(row[8]).strip() if row[8] else '',
            'status': str(row[13]).strip() if row[13] else '',
            'reason': str(reason).strip() if reason not in (None, 'None', '') else '',
        })
    wb.close()

    all_dates = sorted({
        m.group(1) for r in rows
        for m in [re.match(r'(\d{4}-\d{2}-\d{2})', r['time'])]
        if m
    })

    if all_dates:
        latest = all_dates[-1]
        rows = [r for r in rows if r['time'].startswith(latest)]

    return rows, all_dates


def load_watch_list(path: str) -> set:
    if not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    result = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    wb.close()
    return result
