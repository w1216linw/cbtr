import logging
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_log = logging.getLogger('cbt_report')

_HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
_HDR_FONT  = Font(color='FFFFFF', bold=True)
_HDR_ALN   = Alignment(horizontal='center')
_OK_FILL   = PatternFill('solid', fgColor='C6EFCE')
_FL_FILL   = PatternFill('solid', fgColor='FFC7CE')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _save(wb, path: str):
    try:
        wb.save(path)
    except PermissionError:
        msg = f'无法保存 {os.path.basename(path)}：文件被其他程序占用，请关闭后重新运行'
        _log.error(msg)
        sys.exit(f'❌  {msg}')


def pod_date(pod_rows) -> str:
    for r in pod_rows:
        m = re.match(r'(\d{4}-\d{2}-\d{2})', str(r.get('time', '')))
        if m:
            return m.group(1)
    return datetime.today().strftime('%Y-%m-%d')


# ── Daily report ──────────────────────────────────────────────────────────────

def write_report(results, path: str):
    total    = len(results)
    ok_cnt   = sum(1 for r in results if r['status'] == '揽收成功')
    fail_cnt = total - ok_cnt

    failures  = [r for r in results if r['status'] == '揽收失败']
    fail_cell = '\n'.join(f"{r['db_addr']} - {r['reason']}" for r in failures)

    def street_number(addr):
        m = re.match(r'(\d+)', addr.strip())
        return m.group(1) if m else ''

    street_nums = ', '.join(street_number(r['db_addr']) for r in failures if street_number(r['db_addr']))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '揽收报告'

    headers = ['地址库总数', '揽收成功', '揽收失败', '', '揽收失败明细', '揽收失败街道号']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        if not h:
            continue
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _HDR_ALN

    ws.cell(row=2, column=1, value=total)
    ws.cell(row=2, column=2, value=ok_cnt).fill  = _OK_FILL
    ws.cell(row=2, column=3, value=fail_cnt).fill = _FL_FILL
    ws.cell(row=2, column=4, value='')

    e = ws.cell(row=2, column=5, value=fail_cell)
    e.alignment = Alignment(wrap_text=True, vertical='top')
    if failures:
        e.fill = _FL_FILL

    f = ws.cell(row=2, column=6, value=street_nums)
    f.alignment = Alignment(wrap_text=True, vertical='top')
    if failures:
        f.fill = _FL_FILL

    for col, w in zip('ABCDEF', [14, 14, 14, 6, 80, 60]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = max(30, len(failures) * 15)

    _save(wb, path)
