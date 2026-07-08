#!/usr/bin/env python3
"""
CBT Report — Daily Pickup Status Matcher

Files:
  address_db.xlsx       — address list (read-only)
  pod.xlsx              — daily POD (read-only, updated externally)
  report.xlsx           — daily report, OVERWRITTEN each run
  weekly_failures.xlsx  — weekly failed addresses, manually maintained
  address_failures.xlsx — failure count summary, OVERWRITTEN each run
  run.log               — run log (warnings, errors, info)
"""

import logging
import os
import re
import sys
from collections import Counter
from datetime import datetime

import openpyxl

from cbt.loader   import load_db, load_pod, load_watch_list
from cbt.matcher  import build_pod_index, determine_status
from cbt.reporter import (_HDR_ALN, _HDR_FILL, _HDR_FONT, pod_date, write_report)

_log = logging.getLogger('cbt_report')


# ── Logger ────────────────────────────────────────────────────────────────────

def setup_logger(base_dir: str) -> None:
    log_path = os.path.join(base_dir, 'run.log')
    _log.setLevel(logging.DEBUG)
    if _log.handlers:
        _log.handlers.clear()

    fmt = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    fh = logging.FileHandler(log_path, encoding='utf-8', mode='a')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    ch = logging.StreamHandler()
    ch.setLevel(logging.WARNING)
    ch.setFormatter(logging.Formatter('⚠️  [%(levelname)s] %(message)s'))

    _log.addHandler(fh)
    _log.addHandler(ch)

    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f'\n{"=" * 60}\n RUN {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n{"=" * 60}\n')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_header(ws, label: str, width: int = 60):
    ws.column_dimensions['A'].width = width
    c = ws.cell(row=1, column=1, value=label)
    c.fill, c.font, c.alignment = _HDR_FILL, _HDR_FONT, _HDR_ALN


def _extract_addr(raw: str) -> str:
    """Strip the failure reason (Chinese text) from a 'address reason' cell."""
    return re.sub(r'\s*-?\s*\S*[^\x00-\x7F].*$', '', raw).strip()


# ── First-run setup ───────────────────────────────────────────────────────────

def first_run_setup(base_dir: str) -> bool:
    """Create template Excel files if missing. Returns True if any were created."""
    created = []

    addr_path = os.path.join(base_dir, 'address_db.xlsx')
    if not os.path.exists(addr_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '地址库'
        _make_header(ws, '地址')
        ws.cell(row=2, column=1, value='示例：123 Main St, Chicago, IL, 60601（请删除此行，填入实际地址）')
        wb.save(addr_path)
        created.append(('address_db.xlsx', '地址库，每行一个地址'))

    wl_path = os.path.join(base_dir, 'watch_list.xlsx')
    if not os.path.exists(wl_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '监控地址'
        _make_header(ws, '地址')
        wb.save(wl_path)
        created.append(('watch_list.xlsx', '可选，需人工确认的地址'))

    wf_path = os.path.join(base_dir, 'weekly_failures.xlsx')
    if not os.path.exists(wf_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '揽收失败记录'
        _make_header(ws, '地址')
        wb.save(wf_path)
        created.append(('weekly_failures.xlsx', '每周手动填入揽收失败地址，可重复填写'))

    if created:
        print('=' * 56)
        print('  首次运行 — 已自动生成以下模板文件：')
        print('=' * 56)
        for fname, desc in created:
            print(f'  ✅  {fname}  （{desc}）')
        print()
        print('  还需要手动放入同一目录：')
        print('  ❗  pod.xlsx  — 从系统导出的当日揽收数据')
        print()
        print('  文件准备好后，重新运行程序即可。')
        print('=' * 56)
        return True

    return False


# ── Menu ──────────────────────────────────────────────────────────────────────

def show_menu() -> str:
    print()
    print('=' * 46)
    print('  CBT Report')
    print('  1. 每日揽收报告')
    print('  2. 失败地址统计')
    print('=' * 46)
    while True:
        choice = input('  请选择 [1/2]：').strip()
        if choice in ('1', '2'):
            return choice
        print('  请输入 1 或 2')


# ── Feature 1: Daily report ───────────────────────────────────────────────────

def prompt_watch_list_overrides(results: list, watch_list: set) -> bool:
    changed = False
    for r in results:
        if r['status'] != '揽收失败':
            continue
        addr = r['db_addr']
        if addr not in watch_list:
            continue
        print()
        print(f'⚠️  监控地址揽收失败：{addr}')
        while True:
            ans = input('    该地址是否实际取货成功？[y/N] ').strip().lower()
            if ans in ('y', 'yes'):
                r['status'] = '揽收成功'
                r['reason'] = ''
                print('    ✅ 已手动确认为揽收成功。')
                changed = True
                break
            elif ans in ('n', 'no', ''):
                break
            else:
                print('    请输入 y 或 n。')
    return changed


def run_daily_report(base_dir: str):
    db_path     = os.path.join(base_dir, 'address_db.xlsx')
    pod_path    = os.path.join(base_dir, 'pod.xlsx')
    report_path = os.path.join(base_dir, 'report.xlsx')

    # ── 必要文件检查 ─────────────────────────────────────────────────────────
    missing = []
    for path, label in [(db_path, 'address_db.xlsx'), (pod_path, 'pod.xlsx')]:
        if not os.path.exists(path):
            _log.error(f'缺少必要文件：{label}')
            missing.append(label)
    if missing:
        print(f'❌  缺少文件：{", ".join(missing)}，请检查后重新运行')
        input('按 Enter 键退出…')
        return

    # ── 加载数据 ─────────────────────────────────────────────────────────────
    db_addrs            = load_db(db_path)
    pod_rows, pod_dates = load_pod(pod_path)
    watch_list          = load_watch_list(os.path.join(base_dir, 'watch_list.xlsx'))
    date_str            = pod_date(pod_rows)

    _log.info(f'address_db 加载完成：{len(db_addrs)} 条地址')
    _log.info(f'pod.xlsx 加载完成：{len(pod_rows)} 条记录，日期 {date_str}')

    # ── POD 数据校验 ─────────────────────────────────────────────────────────
    if not db_addrs:
        _log.error('address_db.xlsx 中没有地址数据，请检查文件内容')
        print('❌  address_db.xlsx 为空')
        input('按 Enter 键退出…')
        return

    if not pod_rows:
        _log.error('pod.xlsx 中没有有效数据行，请确认文件已正确导出')
        print('❌  pod.xlsx 无有效数据')
        input('按 Enter 键退出…')
        return

    if len(pod_dates) > 1:
        _log.warning(
            f'pod.xlsx 包含多个日期：{", ".join(pod_dates)}；'
            f'已自动使用最新日期 {pod_dates[-1]}，旧日期数据已忽略'
        )

    ratio = len(pod_rows) / len(db_addrs)
    if ratio >= 2:
        _log.warning(
            f'POD 行数（{len(pod_rows)}）是地址库数量（{len(db_addrs)}）的 {ratio:.1f} 倍，'
            f'请确认导出时已按本站点过滤，否则匹配结果可能不准确'
        )

    # ── 匹配 & 输出 ──────────────────────────────────────────────────────────
    full_idx, base_idx = build_pod_index(pod_rows)
    results = [determine_status(addr, full_idx, base_idx) for addr in db_addrs]

    if watch_list:
        overridden = prompt_watch_list_overrides(results, watch_list)
        if overridden:
            print()

    write_report(results, report_path)

    total    = len(results)
    ok_cnt   = sum(1 for r in results if r['status'] == '揽收成功')
    fail_cnt = total - ok_cnt

    _log.info(f'运行完成 — 总地址: {total}  揽收成功: {ok_cnt}  揽收失败: {fail_cnt}')

    print(f'\nPOD日期 : {date_str}')
    print(f'总地址  : {total}  |  揽收成功: {ok_cnt}  |  揽收失败: {fail_cnt}')
    print(f'报告    : {report_path}')
    print()
    print('── 揽收失败明细 ──')
    for r in results:
        if r['status'] == '揽收失败':
            print(f'  [{r["reason"]}]  {r["db_addr"]}')


# ── Feature 2: Failure count summary ─────────────────────────────────────────

def run_summary(base_dir: str):
    input_path  = os.path.join(base_dir, 'weekly_failures.xlsx')
    output_path = os.path.join(base_dir, 'address_failures.xlsx')

    if not os.path.exists(input_path):
        print('❌  找不到 weekly_failures.xlsx，请先创建并填入揽收失败地址。')
        input('按 Enter 键退出…')
        return

    # ── 读取 weekly_failures.xlsx，按地址统计次数 ────────────────────────────
    wb_in = openpyxl.load_workbook(input_path, read_only=True)
    ws_in = wb_in.active
    addrs = []
    for row in ws_in.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            addr = _extract_addr(str(val).strip())
            if addr:
                addrs.append(addr)
    wb_in.close()

    if not addrs:
        print('❌  weekly_failures.xlsx 中没有数据（第2行起填写地址）。')
        input('按 Enter 键退出…')
        return

    counts = Counter(addrs)
    sorted_rows = sorted(counts.items(), key=lambda x: -x[1])

    # ── 写入 address_failures.xlsx ───────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '失败地址统计'
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12

    for col, h in enumerate(['地址', '失败次数'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = _HDR_FILL, _HDR_FONT, _HDR_ALN

    for addr, count in sorted_rows:
        ws.append([addr, count])

    try:
        wb.save(output_path)
    except PermissionError:
        print('❌  无法保存 address_failures.xlsx：文件被占用，请关闭后重试。')
        input('按 Enter 键退出…')
        return

    _log.info(f'失败地址统计完成：{len(counts)} 个地址，共 {len(addrs)} 条记录，输出至 {output_path}')

    print(f'\n读取记录：{len(addrs)} 条  |  独立地址：{len(counts)} 个')
    print(f'输出文件：{output_path}')
    print()
    print('── 失败次数排名 ──')
    for addr, count in sorted_rows[:15]:
        print(f'  {count} 次  {addr}')
    if len(sorted_rows) > 15:
        print(f'  …（共 {len(sorted_rows)} 个，完整列表见 Excel）')


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    setup_logger(base_dir)

    if first_run_setup(base_dir):
        input('按 Enter 键退出…')
        sys.exit(0)

    choice = show_menu()
    print()

    if choice == '1':
        run_daily_report(base_dir)
    else:
        run_summary(base_dir)

    input('\n按 Enter 键退出…')


if __name__ == '__main__':
    main()
